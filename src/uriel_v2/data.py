from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from uriel_v2.models import Draw


ROUND_ALIASES = ("회차", "round", "draw")
NUMBER_ALIASES = tuple((f"번호{i}", f"number{i}", f"num{i}", f"n{i}") for i in range(1, 7))
BONUS_ALIASES = ("보너스", "bonus")


def _normalized(value: object) -> str:
    return str(value).strip().lower().replace(" ", "").replace("_", "")


def _find_column(headers: list[object], aliases: Iterable[str], *, required: bool = True) -> int | None:
    normalized = {_normalized(value): index for index, value in enumerate(headers) if value is not None}
    for alias in aliases:
        index = normalized.get(_normalized(alias))
        if index is not None:
            return index
    if required:
        raise ValueError(f"필수 열을 찾을 수 없습니다: {', '.join(aliases)}")
    return None


def load_draws(path: str | Path, sheet_name: str | None = None) -> list[Draw]:
    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise FileNotFoundError(f"로또 데이터 파일이 없습니다: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"시트를 찾을 수 없습니다: {sheet_name}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        rows = sheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration as exc:
            raise ValueError("엑셀 시트가 비어 있습니다") from exc

        round_column = _find_column(headers, ROUND_ALIASES)
        number_columns = [_find_column(headers, aliases) for aliases in NUMBER_ALIASES]
        bonus_column = _find_column(headers, BONUS_ALIASES, required=False)

        draws: list[Draw] = []
        for excel_row, row in enumerate(rows, start=2):
            if row[round_column] in (None, ""):
                continue
            try:
                round_no = int(row[round_column])
                numbers = tuple(sorted(int(row[column]) for column in number_columns))
                bonus = int(row[bonus_column]) if bonus_column is not None and row[bonus_column] not in (None, "") else None
            except (TypeError, ValueError) as exc:
                raise ValueError(f"{excel_row}행의 회차/번호가 정수가 아닙니다") from exc

            _validate_draw(round_no, numbers, bonus, excel_row)
            draws.append(Draw(round_no=round_no, numbers=numbers, bonus=bonus))
    finally:
        workbook.close()

    if not draws:
        raise ValueError("읽을 수 있는 회차 데이터가 없습니다")

    draws.sort(key=lambda draw: draw.round_no)
    rounds = [draw.round_no for draw in draws]
    if len(rounds) != len(set(rounds)):
        duplicates = sorted({round_no for round_no in rounds if rounds.count(round_no) > 1})
        raise ValueError(f"중복 회차가 있습니다: {duplicates}")
    return draws


def _validate_draw(round_no: int, numbers: tuple[int, ...], bonus: int | None, excel_row: int) -> None:
    if round_no <= 0:
        raise ValueError(f"{excel_row}행의 회차가 올바르지 않습니다: {round_no}")
    if len(numbers) != 6 or len(set(numbers)) != 6:
        raise ValueError(f"{excel_row}행은 서로 다른 당첨번호 6개가 필요합니다: {numbers}")
    if any(number < 1 or number > 45 for number in numbers):
        raise ValueError(f"{excel_row}행의 당첨번호가 1~45 범위를 벗어났습니다: {numbers}")
    if bonus is not None and (bonus < 1 or bonus > 45 or bonus in numbers):
        raise ValueError(f"{excel_row}행의 보너스 번호가 올바르지 않습니다: {bonus}")


def find_draw(draws: Iterable[Draw], round_no: int) -> Draw:
    for draw in draws:
        if draw.round_no == round_no:
            return draw
    raise ValueError(f"회차를 찾을 수 없습니다: {round_no}")
